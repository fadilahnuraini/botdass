import re
import json
import logging
import random
import pickle
import numpy as np
import nltk
import os 
import pandas as pd
import tensorflow as tf
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from sklearn.feature_extraction.text import TfidfVectorizer
from tensorflow.keras.models import load_model
import threading
import asyncio
import torch
from transformers import BertTokenizer, BertForSequenceClassification
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
import logging
import mysql.connector
from datetime import datetime
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import asyncio


def simpan_data_user_ke_excel(user_id, name, age, gender, location,
                               carebot_logs, dass_logs, skor_dass, penilaian_manfaat=None, output_dir="data_excel_user"):

    import os
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)


    wb = Workbook()

    # Sheet Identitas
    identitas_ws = wb.active
    identitas_ws.title = "Identitas"
    identitas_ws.append(["User ID", "Nama", "Umur", "Gender", "Lokasi"])
    identitas_ws.append([user_id, name, age, gender, location])

    # Sheet CareBot
    carebot_ws = wb.create_sheet(title="CareBot Chat")
    carebot_ws.append(["Waktu", "Peran", "Pesan"])
    if carebot_logs:
        for log in carebot_logs:
            carebot_ws.append([log["timestamp"], log["role"], log["message"]])
    else:
        carebot_ws.append(["-", "-", "-"])

    # Sheet Skrining
    skrining_ws = wb.create_sheet(title="Skrining Chat")
    skrining_ws.append(["No", "Waktu", "Pertanyaan", "Jawaban","skor"])
    for idx, log in enumerate(dass_logs, 1):
        skrining_ws.append([idx, log.get("timestamp", ""), log["question"], log["answer"],log.get("score", "-")
        ])

    # Sheet Hasil Skrining
    hasil_ws = wb.create_sheet(title="Hasil Skrining")
    hasil_ws.append(["Aspek", "Skor", "Kategori"])
    for aspek in ["Depresi", "Kecemasan", "Stres"]:
        skor = skor_dass[aspek]["skor"]
        kategori = skor_dass[aspek]["kategori"]
        hasil_ws.append([aspek, skor, kategori])

    # Sheet Rating jika tersedia
    if penilaian_manfaat:
        if penilaian_manfaat.startswith("Rating:"):
            rating_ws = wb.create_sheet(title="Rating")
            rating_ws.append(["Waktu", "Nilai"])
            rating_ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), penilaian_manfaat.replace("Rating: ", "")])
        else:
            manfaat_ws = wb.create_sheet(title="Manfaat Chatbot")
            manfaat_ws.append(["Waktu", "Penilaian"])
            manfaat_ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), penilaian_manfaat])

    # Simpan ke file
    filename = f"{output_dir}/User_{user_id}_{name}.xlsx"
    wb.save(filename)
    

nltk.download('wordnet')
nltk.download('omw-1.4')


# 🔹 Fungsi Simpan Percakapan Chatbot & Skrining

# 🔹 Koneksi ke MySQL
# 🔹 Koneksi ke database MySQL
#conn = mysql.connector.connect(
   # host="localhost",
    #user="root",  
    #password="Anianakke3#",  # Ganti dengan password MySQL Anda
    #database="chatbot_db")
#cursor = conn.cursor()

def save_chat_to_db(user_id, name, age, gender, location, role, message):
    """Menyimpan percakapan ke database MySQL."""
    query = """
    INSERT INTO chat_history (user_id, name, age, gender, location, timestamp, role, message)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """
    values = (user_id, name, age, gender, location, datetime.now(), role, message)

    try:
        cursor.execute(query, values)
        conn.commit()  # Simpan perubahan ke database
        print("✅ Chat berhasil disimpan ke database!")
    except mysql.connector.Error as err:
        print(f"❌ Error MySQL: {err}")

# ===================[ Konfigurasi Logging ]===================

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# Load model BERT untuk analisis jawaban pengguna
MODEL_NAME = "bert-base-uncased"  # Bisa diganti dengan IndoBERT jika pakai bahasa Indonesia
tokenizer = BertTokenizer.from_pretrained(MODEL_NAME)
model = BertForSequenceClassification.from_pretrained(MODEL_NAME, num_labels=4)  # 4 skor DASS-21 (0-3)

# Fungsi untuk mengubah jawaban pengguna menjadi skor DASS-21
def predict_dass21_score(answer):
    inputs = tokenizer(answer, return_tensors="pt", padding=True, truncation=True, max_length=128)
    with torch.no_grad():
        outputs = model(**inputs)
    logits = outputs.logits
    predicted_score = torch.argmax(logits, dim=1).item()  # Pilih skor tertinggi (0,1,2,3)
    return predicted_score

# ===================[ Load Model Chatbot (Neural Network) ]===================
nltk.download('punkt')
lemmatizer = nltk.WordNetLemmatizer()

# Load intents JSON
with open("translated_intents.json", "r", encoding="utf-8") as file:
    intents = json.load(file)

# Load model AI & vectorizer
with open("vectorizer.pkl", "rb") as file:
    vectorizer = pickle.load(file)

with open("label_encoder.pkl", "rb") as file:
    label_encoder = pickle.load(file)

model = load_model("chatbot_model.keras")

def predict_class(text):
    """Memprediksi intent dari input pengguna tanpa mencampur terlalu banyak riwayat"""

    words = nltk.word_tokenize(text)  # Gunakan hanya input terbaru
    processed_text = " ".join([lemmatizer.lemmatize(w.lower()) for w in words])

    X_input = vectorizer.transform([processed_text]).toarray()
    prediction = model.predict(X_input)
    predicted_class = np.argmax(prediction)

    tag = label_encoder.inverse_transform([predicted_class])[0]

    # Cari respons yang sesuai dari intents JSON
    for intent in intents["intents"]:
        if intent["tag"] == tag:
            return random.choice(intent["responses"])

    return "Maaf, saya tidak mengerti. Bisa Anda jelaskan lebih lanjut?"

   


# ===================[ Mapping Keyword ke Skor Skrining ]===================
keywords = {
    0: ["tidak", "biasa saja", "normal", "santai", "tidak pernah", "baik-baik saja", "tidak ada masalah", "saya sangat", "saya merasa lebih baik"],
    1: ["kadang-kadang", "terkadang", "sesekali", "agak terganggu", "lumayan", "sekali-sekali", "terasa sedikit"],
    2: ["sering", "beberapa kali", "cukup berat", "mengganggu", "berulang kali", "sering merasa", "lumayan sering"],
    3: ["sangat sulit", "tidak bisa", "parah", "berat sekali", "sangat mengganggu", "tak tertahankan", "ekstrem"],
}


# Kata negasi yang perlu diperhatikan
negation_words = ["tidak", "bukan", "tidak pernah", "belum", "tidak merasa"]

# ===================[ Fungsi Analisis Jawaban ]===================
def detect_negation(response):
    """Mendeteksi apakah ada kata negasi dalam jawaban pengguna."""
    return any(neg_word in response for neg_word in negation_words)

def analyze_response(response, question):
    """Menganalisis jawaban pengguna dan memberikan skor yang sesuai."""
    response = response.lower().strip()
    detected_scores = set()

    has_negation = detect_negation(response)

    for score in sorted(keywords.keys(), reverse=True):
        if any(re.search(rf"\b{re.escape(word)}\b", response) for word in keywords[score]):
            detected_scores.add(score)

    if has_negation and ("kurang" in question.lower() or "tidak" in question.lower()):
        return 0

    return max(detected_scores, default=0)

# ===================[ Skrining DASS-21 ]===================
user_sessions = {}

dass_questions = [
    {"question": "Dalam beberapa hari terakhir, seberapa sering Anda merasa sulit untuk tenang?", "scale": "stress"},
    {"question": "Seberapa sering Anda merasa mulut Anda kering saat sedang merasa cemas?", "scale": "anxiety"},
    {"question": "Akhir-akhir ini, seberapa sering Anda kesulitan melihat hal-hal positif dalam hidup?", "scale": "depression"},
    {"question": "Dalam beberapa waktu terakhir, seberapa sering Anda merasa tidak bisa bekerja sebaik biasanya?", "scale": "stress"},
    {"question": "Seberapa sering Anda merasa bereaksi berlebihan terhadap situasi yang terjadi?", "scale": "stress"},
    {"question": "Seberapa sering Anda merasa kesulitan bernapas, seperti sesak napas tanpa alasan fisik yang jelas?", "scale": "anxiety"},
    {"question": "Seberapa sering Anda merasa tidak punya dorongan untuk memulai sesuatu?", "scale": "depression"},
    {"question": "Dalam kehidupan sehari-hari, seberapa sering Anda merasa sangat emosional terhadap hal-hal kecil?", "scale": "stress"},
    {"question": "Seberapa sering tangan Anda gemetar tanpa sebab yang jelas?", "scale": "anxiety"},
    {"question": "Seberapa sering Anda merasa tidak tertarik lagi pada hal-hal yang biasanya menyenangkan?", "scale": "depression"},
    {"question": "Dalam beberapa hari terakhir, seberapa sering Anda merasa cemas tanpa tahu kenapa?", "scale": "anxiety"},
    {"question": "Seberapa sering Anda merasa tidak punya harapan tentang masa depan?", "scale": "depression"},
    {"question": "Seberapa sering Anda mudah terganggu oleh hal-hal kecil di sekitar Anda?", "scale": "stress"},
    {"question": "Seberapa sering Anda merasa sulit untuk benar-benar rileks dan tenang?", "scale": "stress"},
    {"question": "Dalam seminggu terakhir, seberapa sering Anda merasa cemas hampir sepanjang waktu?", "scale": "anxiety"},
    {"question": "Seberapa sering Anda merasa sedih atau tertekan belakangan ini?", "scale": "depression"},
    {"question": "Seberapa sering Anda merasa tidak sabaran terhadap hal-hal yang mengganggu Anda?", "scale": "stress"},
    {"question": "Seberapa sering Anda merasa hampir panik meskipun tidak ada alasan yang jelas?", "scale": "anxiety"},
    {"question": "Seberapa sering Anda merasa kehilangan semangat untuk melakukan hal-hal yang biasa Anda sukai?", "scale": "depression"},
    {"question": "Seberapa sering Anda merasa sulit menjaga fokus karena mudah terganggu oleh hal-hal di sekitar?", "scale": "stress"},
    {"question": "Seberapa sering Anda merasa bahwa hidup ini tidak berarti bagi Anda?", "scale": "depression"},
]


dass_interpretation = {
    "depression": [(0, 9, "Normal"), (10, 13, "Ringan"), (14, 20, "Sedang"), (21, 27, "Berat"), (28, 42, "Sangat Berat")],
    "anxiety": [(0, 7, "Normal"), (8, 9, "Ringan"), (10, 14, "Sedang"), (15, 19, "Berat"), (20, 42, "Sangat Berat")],
    "stress": [(0, 14, "Normal"), (15, 18, "Ringan"), (19, 25, "Sedang"), (26, 33, "Berat"), (34, 42, "Sangat Berat")],
}

# ===================[ Fungsi Skrining DASS-21 ]===================


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Menampilkan menu utama dan meminta identitas jika belum ada"""
    user_data = context.user_data

    # ✅ Tampilkan deskripsi fitur terlebih dahulu
    await update.message.reply_text(
        "Saya adalah asisten virtual yang dapat membantu Anda dalam:\n"
        "1️⃣ * Skrining DASS-21 * 🧠 - Tes untuk mengukur tingkat Depresi, Kecemasan, dan Stres.\n"
        "2️⃣ * CareBot * 🤖 - Chatbot yang bisa menjawab pertanyaan Anda tentang kesehatan mental.\n\n"
        "🔹 Ketik `/start` kapan saja untuk kembali ke menu utama.\n"
        "🔹 Pilih salah satu fitur di bawah ini untuk memulai.\n"
        "🔹 Ketik selesai jika ingin mengakhiri chatbot."
    )

    # ✅ Jika identitas belum lengkap, arahkan isi identitas
    if "name" not in user_data or "age" not in user_data or "gender" not in user_data or "location" not in user_data:
        await update.message.reply_text(
            "👋 Selamat datang! Sebelum lanjut, silakan isi identitas Anda terlebih dahulu.",
            reply_markup=ReplyKeyboardRemove()
        )
        await update.message.reply_text("Silakan masukkan nama Anda:")
        context.user_data["awaiting_name"] = True
        return

    # ✅ Jika identitas sudah lengkap, tampilkan menu utama
    await update.message.reply_text("👋 Selamat datang kembali! Anda bisa langsung menggunakan menu di bawah ini:")

    reply_keyboard = [["CareBot", "Skrining DASS-21"]]
    markup = ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True, resize_keyboard=True)

    await update.message.reply_text(
        "Silakan pilih salah satu fitur di bawah ini:",
        reply_markup=markup
    )



async def handle_identity_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Menyimpan identitas pengguna sebelum memulai bot."""
    user_data = context.user_data
    text = update.message.text.strip()

    if "awaiting_name" in user_data:
        user_data["name"] = text
        del user_data["awaiting_name"]
        await update.message.reply_text("Terima kasih! Sekarang masukkan umur Anda:")
        user_data["awaiting_age"] = True

    elif "awaiting_age" in user_data:
        user_data["age"] = text
        del user_data["awaiting_age"]
        await update.message.reply_text("Baik! Sekarang masukkan jenis kelamin Anda (L/P):")
        user_data["awaiting_gender"] = True

    elif "awaiting_gender" in user_data:
        user_data["gender"] = text
        del user_data["awaiting_gender"]
        await update.message.reply_text("Terakhir, masukkan domisili Anda:")
        user_data["awaiting_location"] = True

    elif "awaiting_location" in user_data:
        user_data["location"] = text
        del user_data["awaiting_location"]
        name = user_data.get("name", "Pengguna")
        
        await update.message.reply_text(
            f"✅ Identitas Anda telah disimpan!\n"
            f"Nama: {user_data['name']}\n"
            f"Umur: {user_data['age']}\n"
            f"Jenis Kelamin: {user_data['gender']}\n"
            f"Domisili: {user_data['location']}\n\n"
            ""
        )

        name = user_data["name"]
        await update.message.reply_text(
            f"Hai Sdr. {name}, bagaimana perasaan mu seminggu terakhir ini ? 😊\n"
            f"Langsung saja, kita akan mulai *instrumen skrining DASS-21* untuk mengevaluasi kondisi psikologis Anda.",
            parse_mode="Markdown"
        )

        context.user_data["in_dass"] = True
        await start_dass(update, context)

async def handle_menu_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Menangani pilihan fitur dari menu utama"""
    user_data = context.user_data

    # ✅ Cek apakah identitas sudah lengkap
    if "name" not in user_data or "age" not in user_data or "gender" not in user_data or "location" not in user_data:
        await update.message.reply_text("❗ Anda belum mengisi identitas. Silakan ketik nama Anda untuk memulai pengisian identitas.")
        return

    user_choice = update.message.text

    if user_choice == "CareBot":
        context.user_data["in_dass"] = False
        await update.message.reply_text("Anda memilih CareBot. Ada yang bisa saya bantu? Silakan kirim pertanyaan Anda.", reply_markup=ReplyKeyboardRemove())

    elif user_choice == "Skrining DASS-21":
        context.user_data["in_dass"] = True
        await start_dass(update, context)

async def start_dass(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Memulai sesi skrining DASS-21"""
    user_id = update.effective_user.id
    user_sessions[user_id] = {
        "current_question": 0,
        "scores": {"depression": 0, "anxiety": 0, "stress": 0}
    }
    context.user_data["dass_logs"] = []
    await ask_dass_question(update, context)

async def ask_dass_question(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Menampilkan pertanyaan DASS-21 satu per satu"""
    user_id = update.effective_user.id
    session = user_sessions.get(user_id)

    if not session or session["current_question"] >= len(dass_questions):
        await conclude_dass(update, context)
        return

    # Tampilkan pertanyaan berikutnya
    question_data = dass_questions[session["current_question"]]
    await update.message.reply_text(question_data["question"])


async def handle_dass_response(update: Update, context: ContextTypes.DEFAULT_TYPE): 
    user_id = update.effective_user.id
    session = user_sessions.get(user_id)

    if not session:
        await update.message.reply_text("❗ Sesi skrining tidak ditemukan.")
        return

    # Ambil pertanyaan dan hitung skor
    question = dass_questions[session["current_question"]]
    score = analyze_response(update.message.text, question["question"]) * 2
    session["scores"][question["scale"]] += score

    # Simpan log
    if "dass_logs" not in context.user_data:
        context.user_data["dass_logs"] = []

    context.user_data["dass_logs"].append({
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "question": question["question"],
        "answer": update.message.text,
        "score": score
    })

    # Naikkan indeks pertanyaan
    session["current_question"] += 1

    # Cek apakah pertanyaan terakhir sudah dijawab
    if session["current_question"] >= len(dass_questions):
        await conclude_dass(update, context)
    else:
        await ask_dass_question(update, context)


async def conclude_dass(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Menampilkan hasil akhir skrining dan menyimpannya ke Excel."""
    user_id = update.effective_user.id
    session = user_sessions.pop(user_id)

    if not session:
        await update.message.reply_text("Sesi tidak ditemukan.")
        return

    explanations = {
        "Normal": "Hasil Anda menunjukkan bahwa Anda berada dalam batas normal. Tetap jaga kesehatan mental Anda! 😊",
        "Ringan": "Anda mengalami sedikit gejala, tetapi masih dalam tingkat ringan. Coba luangkan waktu untuk relaksasi dan menjaga keseimbangan hidup. 🧘",
        "Sedang": "Gejala yang Anda alami berada di tingkat sedang. Pertimbangkan untuk berbicara dengan seseorang yang bisa membantu, seperti teman dekat atau keluarga. 💬",
        "Berat": "Hasil menunjukkan tingkat yang cukup tinggi. Mungkin ini saatnya untuk mencari dukungan lebih lanjut, seperti berkonsultasi dengan profesional. 🩺",
        "Sangat Berat": "Tingkat yang Anda alami cukup serius. Sangat disarankan untuk berbicara dengan ahli kesehatan mental atau psikolog. Jangan ragu mencari bantuan. 🤝",
    }

    nama_aspek_map = {
    "depression": "Depresi",
    "anxiety": "Kecemasan",
    "stress": "Stres"
}
    # Simpan ke Excel
    skor_dass = {}
    results = []
    for scale in ["depression", "anxiety", "stress"]:
        total = session["scores"][scale]
        kategori = next(cat for min_v, max_v, cat in dass_interpretation[scale] if min_v <= total <= max_v)
        nama_aspek = nama_aspek_map[scale]
        skor_dass[nama_aspek] = {"skor": total, "kategori": kategori}
        results.append(f"**{nama_aspek}**: {total} ({kategori})\n➡️ {explanations[kategori]}")

    result_text = "\n\n".join(results)

    
    #====================hasil skrining=========================
    context.user_data["skor_dass"] = skor_dass
    carebot_logs = context.user_data.get("carebot_logs", [])
    dass_logs = context.user_data.get("dass_logs", [])
    penilaian_manfaat = context.user_data.get("penilaian_manfaat")
    
    simpan_data_user_ke_excel(
        user_id=user_id,
        name=context.user_data["name"],
        age=context.user_data["age"],
        gender=context.user_data["gender"],
        location=context.user_data["location"],
        carebot_logs=carebot_logs,
        dass_logs=dass_logs,
        skor_dass=skor_dass,
        penilaian_manfaat=penilaian_manfaat
    )

    # Tampilkan hasil ke pengguna
    await update.message.reply_text(
        f"📊 **Hasil Skrining DASS-21 Anda:**\n\n{result_text}\n\n"
        "NB: Ini hanya alat bantu skrining, bukan penegakan diagnosis.\n\nTerima kasih 🙏🏼🙏🏼",
        parse_mode="Markdown"
    )

    context.user_data["in_dass"] = False
    context.user_data["bot_selesai"] = False

    name = context.user_data.get("name", "Pengguna")
    await update.message.reply_text(
        f"🙏 Terima kasih telah menyelesaikan skrining, Sdr. {name}.\n"
        "Sekarang Anda dapat langsung berbicara dengan *CareBot* untuk mendiskusikan apa pun yang Anda rasakan atau butuhkan. 😊",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove()
    )
    if user_id in user_sessions:
        del user_sessions[user_id]

async def save_chat_async(user_id, name, age, gender, location, role, message):
   # """Async function untuk menyimpan chat ke database tanpa menghambat bot."""
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(None, save_chat_to_db, user_id, name, age, gender, location, role, message)


async def handle_chat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_message = update.message.text.strip()

    if context.user_data.get("in_dass", False):
        await handle_dass_response(update, context)
        return

    if "name" not in context.user_data:
        await update.message.reply_text("❗ Anda belum mengisi identitas. Silakan ketik /start untuk mengisi identitas terlebih dahulu.")
        return

    logging.info(f"📩 Pesan diterima dari {user_id}: {user_message}")

    name = context.user_data["name"]
    age = context.user_data["age"]
    gender = context.user_data["gender"]
    location = context.user_data["location"]

    response = predict_class(user_message)
    await update.message.reply_text(response)

    timestamp_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ✅ Koreksi indentasi di sini
    if "carebot_logs" not in context.user_data:
        context.user_data["carebot_logs"] = []

    context.user_data["carebot_logs"].append({"timestamp": timestamp_now, "role": "User", "message": user_message})
    context.user_data["carebot_logs"].append({"timestamp": timestamp_now, "role": "Bot", "message": response})

    asyncio.create_task(save_chat_async(user_id, name, age, gender, location, "User", user_message))
    asyncio.create_task(save_chat_async(user_id, name, age, gender, location, "Bot", response))

    logging.info(f"🤖 Bot membalas: {response}")


async def handle_text_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_data = context.user_data
    text = update.message.text.strip().lower()

    # 🔒 Bot sudah selesai
    if user_data.get("bot_selesai") and text != "/start":
        return

    # Selesai interaksi dengan CareBot
    if text == "selesai" and not user_data.get("in_dass", False):
        await update.message.reply_text(
            "📴 Terima kasih telah berbicara dengan CareBot.\n"
            "Sebelum Anda pergi, kami ingin tahu seberapa bermanfaat chatbot ini bagi Anda.\n\n"
            "Ketik salah satu:\n- Sangat membantu\n- Cukup membantu\n- Kurang membantu\n- Tidak membantu"
        )
        user_data["awaiting_effectiveness"] = True
        return

    # Hentikan skrining
    if text == "berhenti" and user_data.get("in_dass", False):
        user_data["in_dass"] = False
        await update.message.reply_text("🛑 Skrining dihentikan. Silakan ketik /start untuk kembali ke menu.")
        return

    # Isi identitas
    if any(key in user_data for key in ["awaiting_name", "awaiting_age", "awaiting_gender", "awaiting_location"]):
        await handle_identity_input(update, context)
        return

    # Proses skrining
    if user_data.get("in_dass", False):
        await handle_dass_response(update, context)
        return

    # Penilaian efektivitas
    if user_data.get("awaiting_effectiveness"):
        penilaian = text.lower()
        if penilaian in ["sangat membantu", "cukup membantu", "kurang membantu", "tidak membantu"]:
            user_data["penilaian_manfaat"] = penilaian.capitalize()
            user_data["awaiting_effectiveness"] = False
            user_data["bot_selesai"] = True

            # Ambil skor DASS jika sudah ada
            skor_dass = context.user_data.get("skor_dass", {
                "Depresi": {"skor": 0, "kategori": "-"},
                "Kecemasan": {"skor": 0, "kategori": "-"},
                "Stres": {"skor": 0, "kategori": "-"},
            })

            simpan_data_user_ke_excel(
                user_id=update.effective_user.id,
                name=user_data["name"],
                age=user_data["age"],
                gender=user_data["gender"],
                location=user_data["location"],
                carebot_logs=user_data.get("carebot_logs", []),
                dass_logs=user_data.get("dass_logs", []),
                skor_dass=skor_dass,
                penilaian_manfaat=user_data["penilaian_manfaat"]
            )

            await update.message.reply_text(
                "✅ Terima kasih atas penilaian Anda!\n"
                "🔚 Bot telah ditutup. Jika Anda ingin mulai lagi, ketik /start."
            )
            return
        else:
            await update.message.reply_text("Mohon pilih salah satu: Sangat membantu / Cukup membantu / Kurang membantu / Tidak membantu.")
            return

    # Normal chat
    if all(k in user_data for k in ["name", "age", "gender", "location"]):
        await handle_chat(update, context)
        return

    await update.message.reply_text("❗ Anda belum mengisi identitas. Silakan ketik /start untuk memulai.")

   
    
async def reset_chat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Menghapus riwayat percakapan dan memulai dari awal"""
    context.user_data["chat_history"] = []
    await update.message.reply_text("Riwayat percakapan telah dihapus. Silakan mulai percakapan baru.")

from collections import defaultdict

# Penyimpanan semua chat dalam memori (bisa ditingkatkan ke DB)
semua_chat = defaultdict(list)

def simpan_semua_chat(user_id, pengirim, pesan):
    semua_chat[user_id].append({
        "waktu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "pengirim": pengirim,
        "pesan": pesan
    })

# Minta rating setelah skrining selesai
async def minta_rating(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("🙏 Terima kasih telah menggunakan layanan kami. Beri rating dari 1 (buruk) hingga 5 (sangat baik):")
    context.user_data["awaiting_rating"] = True

# Tangani input rating
async def handle_rating_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    rating = update.message.text.strip()
    if rating.isdigit() and 1 <= int(rating) <= 5:
        user_id = update.effective_user.id
        simpan_semua_chat(user_id, "User", f"Rating diberikan: {rating}")
        context.user_data["penilaian_manfaat"] = f"Rating: {rating}"  # ✅ penting!
        await update.message.reply_text("⭐ Terima kasih atas rating Anda!")
        del context.user_data["awaiting_rating"]
    else:
        await update.message.reply_text("Mohon masukkan angka dari 1 sampai 5.")

    simpan_data_user_ke_excel(
    user_id=user_id,
    name=context.user_data["name"],
    age=context.user_data["age"],
    gender=context.user_data["gender"],
    location=context.user_data["location"],
    carebot_logs=context.user_data.get("carebot_logs", []),
    dass_logs=context.user_data.get("dass_logs", []),
    skor_dass={
        "Depresi": {"skor": 0, "kategori": "-"},
        "Kecemasan": {"skor": 0, "kategori": "-"},
        "Stres": {"skor": 0, "kategori": "-"},
    },
    penilaian_manfaat=context.user_data.get("penilaian_manfaat")
)

# ===================[ Konfigurasi Bot Telegram ]===================
application = (
    Application.builder()
    .token("7111956700:AAH3MTcF2UxZuU6JhhaFYYP7UARBlTveWvY")
    .read_timeout(30)  
    .connect_timeout(30)
    .build()
)

# Handler
application.add_handler(CommandHandler("start", start))
application.add_handler(MessageHandler(filters.Regex("^(CareBot|Skrining DASS-21)$"), handle_menu_choice))

application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_input))

application.add_handler(CommandHandler("reset", reset_chat))


# Jalankan bot
async def delete_webhook():
    """Menghapus webhook sebelum menjalankan polling untuk mencegah konflik."""
    await application.bot.delete_webhook(drop_pending_updates=True)

def run_bot():
    """Menjalankan bot dengan event loop yang benar untuk menghindari error asyncio."""
    loop = asyncio.new_event_loop()  # Buat event loop baru
    asyncio.set_event_loop(loop)
    
    # Hapus webhook sebelum polling untuk mencegah konflik
    loop.run_until_complete(delete_webhook())
    
    # Jalankan bot dengan polling
    application.run_polling()

if __name__ == "__main__":
    run_bot()
